import os
import sys
import psycopg2
import pandas as pd
import logging
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- CONFIGURATION ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

load_dotenv(dotenv_path=os.path.join(os.path.dirname(BASE_DIR), "config", "db.env"))

def get_pg_conn():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=int(os.getenv("DB_PORT", "5432"))
    )

def run_exceptions_report(target_date_str):
    try:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
        formatted_date = target_date.strftime('%Y-%m-%d')
        
        template_path = "templates/template_stock_exceptions.xlsx"
        output_path = f"outputs/Stock_Exceptions_{formatted_date}.xlsx"

        if not os.path.exists(template_path):
            logger.error(f"❌ Template absent : {template_path}")
            return

        wb = load_workbook(template_path)
        ws = wb.active

        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row + 1)

        conn = get_pg_conn()
        with open("queries/q07_stock_exceptions.sql", "r") as f:
            query = f.read()

        logger.info(f"⚠️ Extraction Exceptions Stock pour le {formatted_date}...")
        df = pd.read_sql_query(query, conn, params={"target_date": formatted_date})
        
        if not df.empty:
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            logger.info(f"✅ {len(df)} anomalies détectées et enregistrées.")
        else:
            logger.info(f"🟢 Aucune exception pour le {formatted_date}. Base propre !")

        conn.close()
        wb.save(output_path)
        logger.info(f"✨ Fichier généré : {output_path}")

    except Exception as e:
        logger.error(f"💥 ERREUR : {str(e)}")

if __name__ == "__main__":
    date_arg = sys.argv[1] if len(sys.argv) > 1 else datetime.now().strftime('%Y-%m-%d')
    run_exceptions_report(date_arg)
