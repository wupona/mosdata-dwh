import os
import sys
import psycopg2
import pandas as pd
import logging
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- RÉPERTOIRE PRINCIPAL ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

# Configuration Verbose (Correction pour garantir l'affichage console)
logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# Chargement config DB
dotenv_path = os.path.join(os.path.dirname(BASE_DIR), "config", "db.env")
load_dotenv(dotenv_path=dotenv_path)

def get_pg_conn():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=int(os.getenv("DB_PORT", "5432"))
    )

def run_monthly_rebuild(target_date_str):
    try:
        # target_date = Jour J (Aujourd'hui)
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
        start_date = target_date.replace(day=1)
        month_suffix = target_date.strftime('%Y_%m')
        
        # On traite du 1er du mois jusqu'au jour J inclus
        # Pour chaque jour 'd', on prendra les mouvements de 'd-1'
        date_list = [start_date + timedelta(days=x) for x in range((target_date - start_date).days + 1)]
        
        logger.info(f"🚀 RECONSTRUCTION : DU {date_list[0].date()} AU {date_list[-1].date()}")
    except Exception as e:
        logger.error(f"❌ Erreur init : {e}")
        return

    template_path = "templates/template_situation_stock.xlsx"
    output_path = f"outputs/Situation_Stock_{month_suffix}.xlsx"

    if not os.path.exists(template_path):
        logger.error(f"❌ Template absent : {template_path}")
        return

    wb = load_workbook(template_path)
    
    # Nettoyage des feuilles
    for s_name in ["Situation_Stock", "Mouvement_Stock", "Age_Stock"]:
        if s_name in wb.sheetnames:
            ws = wb[s_name]
            if ws.max_row >= 2:
                ws.delete_rows(2, ws.max_row + 1)
            logger.info(f"🧹 Feuille '{s_name}' vidée.")

    try:
        conn = get_pg_conn()
        
        for current_day in date_list:
            # LOGIQUE : 
            # Si current_day = 2026-02-01
            # d_j_moins_1 = 2026-01-31 (Mouvements de la veille)
            # d_j         = 2026-02-01 (Ouverture ce matin)
            
            d_j_moins_1 = (current_day - timedelta(days=1)).strftime('%Y-%m-%d')
            d_j = current_day.strftime('%Y-%m-%d')
            
            logger.info(f"📅 --- TRAITEMENT J: {d_j} (via Mvts J-1: {d_j_moins_1}) ---")

            configs = [
                {
                    "label": "SITUATION",
                    "sheet": "Situation_Stock", 
                    "sql": "queries/q02_situation_stock.sql", 
                    "params": {"opening_prev": d_j_moins_1, "movement_day": d_j_moins_1, "opening_curr": d_j}
                },
                {
                    "label": "MOUVEMENT",
                    "sheet": "Mouvement_Stock", 
                    "sql": "queries/q03_mouvement_stock.sql", 
                    "params": {"movement_day": d_j_moins_1}
                },
                {
                    "label": "AGE",
                    "sheet": "Age_Stock", 
                    "sql": "queries/q04_age_stock.sql", 
                    "params": {"opening_curr": d_j}
                }
            ]

            for cfg in configs:
                sql_path = cfg["sql"]
                if not os.path.exists(sql_path): continue

                with open(sql_path, "r") as f:
                    query = f.read()
                
                df = pd.read_sql_query(query, conn, params=cfg["params"])
                
                if df.empty:
                    logger.info(f"  [-] {cfg['label']}: Aucune donnée.")
                    continue

                ws = wb[cfg["sheet"]]
                start_row = ws.max_row + 1 if ws.max_row >= 2 else 2

                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                    for c_idx, value in enumerate(row, start=1):
                        ws.cell(row=r_idx, column=c_idx, value=value)
                
                logger.info(f"  [OK] {cfg['label']}: {len(df)} lignes ajoutées.")

        conn.close()
        wb.save(output_path)
        logger.info(f"✨ TERMINÉ : {output_path}")

    except Exception as e:
        logger.error(f"💥 ERREUR : {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        run_monthly_rebuild(sys.argv[1])
    else:
        print("\nPrécisez une date : python3 r02_situation_stock.py 2026-02-01\n")
