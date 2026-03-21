import os
import sys
import psycopg2
import pandas as pd
import logging
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- DÉCLARATION DU RÉPERTOIRE PRINCIPAL ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

# Configuration du Logging Verbose
logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# Chargement de la config DB
dotenv_path = os.path.join(os.path.dirname(BASE_DIR), "config", "db.env")
load_dotenv(dotenv_path=dotenv_path)

def get_pg_conn():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=int(os.getenv("DB_PORT", "5432"))
    )

def run_detailed_revenue_report(target_date_str):
    try:
        # target_date = Jour J (Aujourd'hui)
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
        start_date = target_date.replace(day=1)
        month_suffix = target_date.strftime('%Y_%m')
        
        # Pour le revenu, on traite généralement du 1er au jour J inclus
        date_list = [start_date + timedelta(days=x) for x in range((target_date - start_date).days + 1)]
        
        logger.info(f"🚀 REVENU DÉTAILLÉ : RECONSTRUCTION DU {start_date.date()} AU {target_date.date()}")
    except Exception as e:
        logger.error(f"❌ Erreur initialisation : {e}")
        return

    template_path = "templates/template_vente_detaillee.xlsx"
    output_path = f"outputs/Vente_Detaillee_{month_suffix}.xlsx"

    if not os.path.exists(template_path):
        logger.error(f"❌ Template absent : {template_path}")
        return

    wb = load_workbook(template_path)
    sheet_name = "vente_detaillee"

    # Nettoyage de la feuille unique
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row >= 2:
            ws.delete_rows(2, ws.max_row + 1)
        logger.info(f"🧹 Feuille '{sheet_name}' réinitialisée.")
    else:
        logger.error(f"❌ La feuille '{sheet_name}' n'existe pas dans le template.")
        return

    try:
        conn = get_pg_conn()
        sql_file = "queries/q05_detailed_revenue.sql"
        
        if not os.path.exists(sql_file):
            logger.error(f"❌ Fichier SQL introuvable : {sql_file}")
            return

        with open(sql_file, "r") as f:
            query = f.read()

        for current_day in date_list:
            day_str = current_day.strftime('%Y-%m-%d')
            
            # Exécution de la requête pour le jour donné
            df = pd.read_sql_query(query, conn, params={"target_date": day_str})
            
            if df.empty:
                logger.info(f"📅 {day_str} : Aucune vente enregistrée.")
                continue

            ws = wb[sheet_name]
            # Déterminer la ligne de départ (A2 au début, puis à la suite)
            start_row = ws.max_row + 1 if ws.max_row >= 2 else 2

            # Insertion des données
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            logger.info(f"📅 {day_str} : [OK] +{len(df)} lignes ajoutées.")

        conn.close()
        wb.save(output_path)
        logger.info(f"✨ RAPPORT GÉNÉRÉ : {output_path}")

    except Exception as e:
        logger.error(f"💥 ERREUR : {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        run_detailed_revenue_report(sys.argv[1])
    else:
        print("\nUsage: python3 r05_detailed_revenue.py 2026-02-09\n")
