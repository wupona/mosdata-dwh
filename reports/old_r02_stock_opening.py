import os
import sys
import psycopg2
import pandas as pd
import logging
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- DÉCLARATION DU RÉPERTOIRE PRINCIPAL ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

# Configuration du Logging Verbose
logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# Chargement de la config DB (un cran au dessus de reports/)
dotenv_path = os.path.join(os.path.dirname(BASE_DIR), "config", "db.env")
load_dotenv(dotenv_path=dotenv_path)

def get_pg_conn():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=int(os.getenv("DB_PORT", "5432"))
    )

def run_stock_opening_report(target_date_str):
    try:
        # target_date = Jour J
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
        start_date = target_date.replace(day=1)
        month_suffix = target_date.strftime('%Y_%m')
        
        # Liste des jours du 1er au jour J inclus
        date_list = [start_date + timedelta(days=x) for x in range((target_date - start_date).days + 1)]
        
        logger.info(f"🚀 STOCK OPENING : RECONSTRUCTION DU MOIS {target_date.strftime('%B %Y')}")
    except Exception as e:
        logger.error(f"❌ Erreur initialisation : {e}")
        return

    template_path = "templates/template_stock_opening.xlsx"
    output_path = f"outputs/Stock_Opening_{month_suffix}.xlsx"

    if not os.path.exists(template_path):
        logger.error(f"❌ Template absent : {template_path}")
        return

    wb = load_workbook(template_path)

    try:
        conn = get_pg_conn()
        sql_file = "queries/q06_stock_opening.sql"
        
        if not os.path.exists(sql_file):
            logger.error(f"❌ Fichier SQL introuvable : {sql_file}")
            return

        with open(sql_file, "r") as f:
            query = f.read()

        for current_day in date_list:
            day_str = current_day.strftime('%Y-%m-%d')
            sheet_name = current_day.strftime('%d') # "01", "02", ...
            
            logger.info(f"📅 Traitement du {day_str} -> Feuille '{sheet_name}'")

            # Vérification de l'existence de la feuille dans le template
            if sheet_name not in wb.sheetnames:
                logger.warning(f"  [!] Feuille '{sheet_name}' absente du template. Passage au jour suivant.")
                continue

            ws = wb[sheet_name]
            
            # Nettoyage de la feuille (A2 à la fin)
            if ws.max_row >= 2:
                ws.delete_rows(2, ws.max_row + 1)

            # Exécution de la requête
            df = pd.read_sql_query(query, conn, params={"target_date": day_str})
            
            if df.empty:
                logger.info(f"  [-] Aucun stock d'ouverture pour ce jour.")
                continue

            # Insertion des données en A2
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            logger.info(f"  [OK] +{len(df)} lignes insérées dans '{sheet_name}'.")

        conn.close()
        wb.save(output_path)
        logger.info(f"✨ RAPPORT GÉNÉRÉ : {output_path}")

    except Exception as e:
        logger.error(f"💥 ERREUR : {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        run_stock_opening_report(sys.argv[1])
    else:
        print("\nUsage: python3 r06_stock_opening.py 2026-02-09\n")
