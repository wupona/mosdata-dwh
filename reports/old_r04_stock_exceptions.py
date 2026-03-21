import os
import sys
import psycopg2
import pandas as pd
import logging
from datetime import datetime, timedelta
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# --- CONFIGURATION ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
os.chdir(BASE_DIR)

logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

load_dotenv(dotenv_path=os.path.join(os.path.dirname(BASE_DIR), "config", "db.env"))

def get_pg_conn():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=int(os.getenv("DB_PORT", "5432"))
    )

def run_exceptions_report(target_date_str):
    try:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d')
        start_date = target_date.replace(day=1)
        month_suffix = target_date.strftime('%Y_%m')
        date_list = [start_date + timedelta(days=x) for x in range((target_date - start_date).days + 1)]
        
        logger.info(f"⚠️ EXCEPTIONS STOCK : VENTILATION PAR ONGLET JUSQU'AU {target_date.date()}")
    except Exception as e:
        logger.error(f"❌ Erreur init : {e}")
        return

    template_path = "templates/template_stock_exceptions.xlsx"
    output_path = f"outputs/Exceptions_Stock_{month_suffix}.xlsx"

    if not os.path.exists(template_path):
        logger.error(f"❌ Template absent : {template_path}")
        return

    wb = load_workbook(template_path)

    try:
        conn = get_pg_conn()
        with open("queries/q07_stock_exceptions.sql", "r") as f:
            query = f.read()

        for current_day in date_list:
            day_str = current_day.strftime('%Y-%m-%d')
            sheet_name = current_day.strftime('%d')
            
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            
            # Nettoyage de la feuille (A2 à la fin)
            if ws.max_row >= 2:
                ws.delete_rows(2, ws.max_row + 1)

            df = pd.read_sql_query(query, conn, params={"target_date": day_str})
            
            if df.empty:
                logger.info(f"📅 {day_str} (Feuille {sheet_name}) : 0 exception.")
                continue

            # Insertion des données en A2
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
                for c_idx, value in enumerate(row, start=1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
            
            logger.info(f"📅 {day_str} (Feuille {sheet_name}) : {len(df)} anomalies trouvées.")

        conn.close()
        wb.save(output_path)
        logger.info(f"✨ RAPPORT GÉNÉRÉ : {output_path}")

    except Exception as e:
        logger.error(f"💥 ERREUR : {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        run_exceptions_report(sys.argv[1])
    else:
        print("\nUsage: python3 r07_stock_exceptions.py 2026-02-09\n")
