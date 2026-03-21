import os
import sys
import logging
from datetime import datetime, date, timedelta
from pathlib import Path

import psycopg2
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

# ---------------------------------------------------------------------
# Rapport Encaissement Journalier (Excel)
# - Remplit TOUTES les feuilles du mois, du 1er jour jusqu'à la date cible
# - La date cible est passée en paramètre: AAAA-MM-JJ
# ---------------------------------------------------------------------

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

REPORTS_ROOT = Path("/mnt/c/Blissydah/reports").resolve()
OUTPUT_DIR = REPORTS_ROOT / "outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)



BASE_DIR = Path(__file__).resolve().parent

# Chargement config DB
# Par défaut: ./config/db.env (à côté du script)
env_path = BASE_DIR / "config" / "db.env"
if env_path.exists():
    load_dotenv(dotenv_path=env_path)
else:
    # Fallback: si vous exécutez depuis un autre répertoire avec config/db.env
    load_dotenv(dotenv_path=Path("config") / "db.env")


def get_pg_conn():
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=int(os.getenv("DB_PORT", "5432"))
    )


def daterange(d1: date, d2: date):
    """Yield each day from d1 to d2 inclusive."""
    cur = d1
    while cur <= d2:
        yield cur
        cur += timedelta(days=1)


def clear_sheet_area(ws, min_row=4, max_row=50, min_col=1, max_col=12):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None


def write_df_to_sheet(ws, df: pd.DataFrame, start_row=4, start_col=1):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)


def run_report(target_date_str: str):
    # 0) Paramètre date
    try:
        target_date = datetime.strptime(target_date_str, '%Y-%m-%d').date()
    except ValueError:
        logger.error("Format date invalide. Attendu: AAAA-MM-JJ (ex: 2026-02-11)")
        return 2

    month_suffix = target_date.strftime('%Y_%m')  # ex: 2026_02

    # 1) Chemins (relatifs au script, avec fallback)
    template_path = BASE_DIR / "reports" / "templates" / "template_rapport_encaissement.xlsx"
    if not template_path.exists():
        template_path = Path("reports/templates/template_rapport_encaissement.xlsx")

    query_path = BASE_DIR / "reports" / "queries" / "q01_daily_revenue.sql"
    if not query_path.exists():
        query_path = Path("reports/queries/q01_daily_revenue.sql")

    ##output_dir = BASE_DIR / "reports" / "outputs"
    ##output_dir.mkdir(parents=True, exist_ok=True)

    output_file = OUTPUT_DIR / f"Rapport_Encaissement_{month_suffix}.xlsx"

    # 2) Source du classeur (fichier du mois existant OU template)
    if output_file.exists():
        source_path = output_file
        logger.info(f"Mise à jour du fichier mensuel existant : {output_file}")
    else:
        source_path = template_path
        logger.info(f"Nouveau mois détecté. Création du fichier à partir du template : {output_file}")

    if not Path(source_path).exists():
        logger.error(f"Template / source introuvable: {source_path}")
        return 2

    if not query_path.exists():
        logger.error(f"Query SQL introuvable: {query_path}")
        return 2

    query = query_path.read_text(encoding="utf-8")

    # 3) Charger le workbook une seule fois
    wb = load_workbook(source_path)

    # 4) DB: une seule connexion pour tout le mois
    conn = None
    try:
        conn = get_pg_conn()

        start_date = target_date.replace(day=1)

        for d in daterange(start_date, target_date):
            sheet_name = d.strftime('%d')  # '01', '02', ...

            if sheet_name not in wb.sheetnames:
                logger.warning(f"Feuille {sheet_name} introuvable (date={d.isoformat()}). Ignorée.")
                continue

            df = pd.read_sql_query(query, conn, params={"target_date": d.isoformat()})

            ws = wb[sheet_name]

            # Nettoyage préventif (évite résidus si relance)
            clear_sheet_area(ws, min_row=4, max_row=50, min_col=1, max_col=12)

            if df.empty:
                logger.warning(f"Aucune donnée pour le {d.isoformat()} (feuille {sheet_name}). Feuille vidée.")
                continue

            write_df_to_sheet(ws, df, start_row=4, start_col=1)
            logger.info(f"OK: feuille {sheet_name} remplie ({len(df)} lignes)")

        # 5) Sauvegarde finale
        wb.save(output_file)
        logger.info(f"✅ Succès : fichier mis à jour => {output_file}")
        return 0

    except Exception as e:
        logger.error(f"Erreur : {e}", exc_info=True)
        return 2
    finally:
        try:
            if conn is not None:
                conn.close()
        except Exception:
            pass


if __name__ == "__main__":
    if len(sys.argv) > 1:
        sys.exit(run_report(sys.argv[1]))
    else:
        logger.error("Usage: python3 r01_encaissement_journalier.py AAAA-MM-JJ")
        sys.exit(2)
