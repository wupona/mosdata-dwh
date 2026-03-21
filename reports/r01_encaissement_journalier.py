#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
r01_encaissement_journalier.py
------------------------------------------------------------
Rapport Encaissement journalier (feuille unique) - REBUILD du mois jusqu'au Jour J

Objectif
- Le rapport "Encaissement journalier" doit TOUJOURS contenir tous les jours du mois,
  du 1er jusqu’au jour J, même si le script n’a pas tourné certains jours.

Comportement
- 1 seule date (Jour J)  => reconstruit du 1er du mois jusqu’au Jour J (inclus)
- 2 dates               => reconstruit la plage [date_from ; date_to] (inclusive)

Usage
  python r01_encaissement_journalier.py YYYY-MM-DD
  python r01_encaissement_journalier.py YYYY-MM-DD YYYY-MM-DD
------------------------------------------------------------
"""

import os
import sys
import logging
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Optional, Tuple, Iterator, List

import psycopg2
import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------
# Logging
# ---------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger("r01_encaissement")


# ---------------------------
# Paths robustes (root projet)
# ---------------------------
SCRIPT_DIR = Path(__file__).resolve().parent


def find_project_root(start: Path) -> Path:
    """
    Remonte depuis le dossier du script jusqu'à trouver un dossier contenant:
    - config/
    - reports/
    """
    for p in [start, *start.parents]:
        if (p / "config").is_dir() and (p / "reports").is_dir():
            return p
    return start  # fallback


ROOT_DIR = find_project_root(SCRIPT_DIR)


# ---------------------------
# Helpers dates
# ---------------------------
def parse_iso_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def iter_days(d_from: date, d_to: date) -> Iterator[date]:
    cur = d_from
    while cur <= d_to:
        yield cur
        cur += timedelta(days=1)


def compute_period(d_from: date, d_to: Optional[date]) -> Tuple[date, date]:
    """
    Règle métier:
    - si d_to est None => [1er du mois de d_from ; d_from] inclus
    - sinon => [d_from ; d_to] inclus
    """
    if d_to is None:
        return d_from.replace(day=1), d_from

    if d_to < d_from:
        raise ValueError("date_to < date_from")

    return d_from, d_to


# ---------------------------
# DB
# ---------------------------
def load_env() -> None:
    env_path = ROOT_DIR / "config" / "db.env"
    if not env_path.exists():
        logger.error(f"db.env introuvable: {env_path}")
        sys.exit(2)
    load_dotenv(dotenv_path=str(env_path))


def get_pg_conn():
    """
    Connexion PostgreSQL via variables d'environnement
    Attend:
      DB_HOST, DB_NAME, DB_USER, DB_PASSWORD, DB_PORT
    """
    return psycopg2.connect(
        host=os.getenv("DB_HOST"),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        port=int(os.getenv("DB_PORT", "5432")),
    )


# ---------------------------
# IO
# ---------------------------
def load_sql_query(sql_path: Path) -> str:
    if not sql_path.exists():
        raise FileNotFoundError(f"SQL introuvable: {sql_path}")
    return sql_path.read_text(encoding="utf-8")


def clear_sheet_data(ws, row_start: int, col_start: int, col_end: int, max_rows: int = 10000) -> None:
    """
    Nettoie une zone data de la feuille pour éviter les résidus.
    """
    for r in range(row_start, row_start + max_rows):
        for c in range(col_start, col_end + 1):
            ws.cell(row=r, column=c).value = None


# ---------------------------
# Main logique
# ---------------------------
def run_report(date_from_str: str, date_to_str: Optional[str] = None) -> int:
    # 1) Dates
    try:
        d_from = parse_iso_date(date_from_str)
        d_to = parse_iso_date(date_to_str) if date_to_str else None
        period_from, period_to = compute_period(d_from, d_to)
    except Exception:
        logger.error("Format date invalide. Usage: YYYY-MM-DD [YYYY-MM-DD]")
        return 2

    logger.info(f"🔁 REBUILD ENCAISSEMENT: {period_from.isoformat()} -> {period_to.isoformat()} (inclus)")

    # 2) Chemins (template, output, SQL)
    month_suffix = period_to.strftime("%Y_%m")

    # Template: variable d'env RA_TEMPLATE_PATH sinon path standard sous ROOT
    ra_tpl_env = os.getenv("RA_TEMPLATE_PATH")
    if ra_tpl_env:
        template_path = Path(ra_tpl_env)
    else:
        template_path = ROOT_DIR / "reports" / "templates" / "template_rapport_encaissement.xlsx"

    if not template_path.exists():
        logger.error(f"Template introuvable : {template_path}")
        return 2

    output_dir = ROOT_DIR / "reports" / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_dir / f"Rapport_Encaissement_{month_suffix}.xlsx"

    # Source: si le fichier mensuel existe, on le met à jour; sinon, on part du template
    if output_file.exists():
        source_path = output_file
        logger.info(f"📌 Mise à jour du fichier mensuel existant : {output_file}")
    else:
        source_path = template_path
        logger.info(f"📌 Nouveau mois détecté. Création du fichier à partir du template : {output_file}")

    sql_path = ROOT_DIR / "reports" / "queries" / "q01_daily_revenue.sql"
    try:
        query = load_sql_query(sql_path)
    except Exception as e:
        logger.error(f"Erreur lecture SQL: {sql_path} | {e}")
        return 2

    # 3) Ouvrir workbook
    try:
        wb = load_workbook(str(source_path))
    except Exception as e:
        logger.error(f"Impossible d'ouvrir le fichier source: {source_path} | {e}")
        return 2

    sheet_name = "Encaissement"
    if sheet_name not in wb.sheetnames:
        logger.error(f"Feuille '{sheet_name}' introuvable. Feuilles: {wb.sheetnames}")
        return 2

    ws = wb[sheet_name]

    # 4) Extraction + concat
    all_dfs: List[pd.DataFrame] = []
    conn = None
    try:
        conn = get_pg_conn()

        for d in iter_days(period_from, period_to):
            d_str = d.isoformat()
            logger.info(f"Extraction encaissement: {d_str}")

            try:
                df = pd.read_sql_query(query, conn, params={"target_date": d_str})
            except Exception as e:
                logger.error(f"Erreur SQL pour {d_str}: {e}")
                continue

            if df is None or df.empty:
                logger.warning(f"Aucune donnée pour le {d_str}")
                continue

            # Si votre SQL renvoie déjà une colonne date, on ne force pas.
            # Sinon, on peut en ajouter une colonne standard "period_date".
            if "period_date" not in df.columns and "date" not in df.columns and "Date" not in df.columns:
                df["period_date"] = d_str

            all_dfs.append(df)

    finally:
        if conn is not None:
            conn.close()

    # 5) Ecriture feuille unique (rebuild)
    if not all_dfs:
        # On vide quand même la feuille pour refléter qu'il n'y a rien sur la période
        clear_sheet_data(ws, row_start=4, col_start=1, col_end=30, max_rows=10000)
        try:
            wb.save(str(output_file))
            logger.info(f"✅ Fichier sauvegardé (sans données) : {output_file}")
        except Exception as e:
            logger.error(f"Erreur sauvegarde: {output_file} | {e}")
            return 2
        return 0

    final_df = pd.concat(all_dfs, ignore_index=True)

    # Si "Numero" existe, on le recalcul (1..n) pour l’affichage
    if "Numero" in final_df.columns:
        final_df["Numero"] = range(1, len(final_df) + 1)

    # Clear zone data (col_end dynamique, mais on garde un minimum)
    col_end = max(len(final_df.columns), 12)
    col_end = min(col_end, 50)  # sécurité
    clear_sheet_data(ws, row_start=4, col_start=1, col_end=col_end, max_rows=10000)

    # Write à partir de la ligne 4 (comme vos templates)
    start_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            if c_idx <= col_end:
                ws.cell(row=r_idx, column=c_idx, value=value)

    # 6) Sauvegarde
    try:
        wb.save(str(output_file))
        logger.info(f"✅ Succès : Feuille '{sheet_name}' reconstruite dans {output_file}")
    except Exception as e:
        logger.error(f"Erreur sauvegarde: {output_file} | {e}")
        return 2

    return 0


def main() -> None:
    load_env()

    if len(sys.argv) < 2:
        logger.error("Usage: python r01_encaissement_journalier.py YYYY-MM-DD [YYYY-MM-DD]")
        sys.exit(2)

    date_from = sys.argv[1]
    date_to = sys.argv[2] if len(sys.argv) >= 3 else None

    rc = run_report(date_from, date_to)
    sys.exit(rc)


if __name__ == "__main__":
    main()